# Python Code for Data Analysis and Visualization with pandas, openpyxl, and Matplotlib
import pandas as pd
from openpyxl import load_workbook
from itertools import islice
import matplotlib.pyplot as plt


from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Series, Reference

# Creating a Function to load each excel sheet into a dataframe
def function(origin,datafile):
    wb = load_workbook(filename = origin)
    ws = wb [datafile]
    dataframe = pd.DataFrame(ws.values)
    return dataframe

# Applying the function 
df_metadata = function('week 9/P11-Country-Metadata (1).xlsx', 'Metadata - Countries')
df_fertility = function('week 9/P11-Fertility-Rate.xlsx', 'Data')
df_life_expectancy = function('week 9/P11-Life-Expectancy-At-Birth.xlsx', 'Data' )


# Displaying the first few rows of each DataFrame for quick inspection
# print(df_metadata.head())
# print("==" * 50)
# print(df_fertility.head())
# print("==" * 50)
# print(df_life_expectancy.head())
# print("==" * 50)

#Function to Create worksheets from my workbooks
def Excel(workbooks, worksheets):
    wbb = load_workbook(filename = workbooks)
    wss = wbb [worksheets]
    exl_worksheet = pd.read_excel(workbooks)
    return exl_worksheet

df_metadata3 = Excel('week 9/P11-Country-Metadata (1).xlsx','Metadata - Countries')
df_fertility3 = Excel('week 9/P11-Fertility-Rate.xlsx','Data')
df_life_expectancy3 = Excel( 'week 9/P11-Life-Expectancy-At-Birth.xlsx','Data',)

#print(df_metadata3.head())
#print(df_fertility3.head())
#print(df_life_expectancy3.head())



#   Melting the Year and value(statistic) for Fertility and Life_Expectance datasets
df_cols = ["Country Name", "Country Code", "Indicator Name", "Indicator Code"]
melted_fertility = pd.melt(df_fertility3, id_vars= df_cols, var_name="Year", value_name = "Fertility")


df_cols = ["Country Name", "Country Code", "Indicator Name", "Indicator Code"]
melted_life_expectancy = pd.melt(df_life_expectancy3, id_vars= df_cols, var_name="Year", value_name = "Life Expectancy")


#print(melted_fertility)
#print(melted_life_expectancy)

# Verifying the datatypes of the clolumns in the the dataset
#print(melted_fertility.info())
#print(melted_life_expectancy.info())



# Converting the datatype of the Year column from Object to Integer for both fertility and life_expectancy datasets
melted_fertility ['Year'] = melted_fertility ['Year'].astype(int)
melted_life_expectancy ['Year'] = melted_life_expectancy ['Year'].astype(int)

#print(melted_fertility.info())
#print(melted_life_expectancy.info())


# To display the maximum number of columns to display when showing a DataFrame.
pd.set_option('display.max_columns',500)


# Merging the first two datasets (fertility and metadata) together
df_first_merge = pd.merge (left=df_metadata3,right=melted_fertility,how='left',on=['Country Code'],suffixes = ['L', 'R'])
#print(df_first_merge)


# Merging the the newly comined datasets (fertility and metadata) with life_expectancy
df_second_merge = pd.merge (left=df_first_merge,right=melted_life_expectancy,how='left',on=['Country Code'],suffixes = ['L', 'R'])
#print(df_second_merge)

#Analysis
# Creating a dataframe for my StudentID (This filters and shows only the records applicable to 100871852)
df_studentid = df_second_merge[df_second_merge['StudentID'] == 100871852]

#print(df_studentid)
#print(df_studentid.info())


#Average fertility and life expectancy rate for Swaziland over the years
df_country = df_studentid.groupby(['Country Name'])[['Fertility', 'Life Expectancy']].mean().reset_index()
#print(df_country)


# Grouping the data by region to analyze fertility and life expectancy trends(groupby)
df_region = df_second_merge.groupby(['Region'])[['Fertility','Life Expectancy']].mean().reset_index()
df_region = df_region.iloc[[6]]
# print(df_region)

# Creating a common column name for both dataframes (Region and country) so as to perform a join 
df_country.columns = ['Location', 'Fertility', 'Life Expectancy']
df_region.columns = ['Location', 'Fertility', 'Life Expectancy'] 

# print(df_region)
# print(df_country)

# Question 1
# Joining the dataframes region and country together to create a 
df_join = pd.concat ([df_region, df_country], axis=0)
# print(df_join)


# # Creating a new Excel workbook and active worksheet
wb = Workbook()                 
ws = wb.active

# # Writing data from a pandas DataFrame to the Excel worksheet
# for r in dataframe_to_rows(df_join, index=False, header=True):
#     ws.append(r)

# # Bar chart showing average fertility statistic of Swaziland compared to Sub-saharan region
# last_row = len(ws['A'])
# chart1 = BarChart()
# chart1.type = "col"
# chart1.style = 10
# chart1.title = "Average Fertility Rate"
# chart1.y_axis.title = 'Avg Fertility'
# chart1.x_axis.title = 'Category'

# data = Reference(ws, min_col=2, min_row=1, max_row=last_row, max_col=2)
# cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
# chart1.add_data(data, titles_from_data=True)
# chart1.set_categories(cats)
# chart1.shape = 4
# ws.add_chart(chart1, "E3")

# wb.save('Bar chart showing average fertility statistic of Swaziland compared to Sub-saharan region.xlsx')


# # Bar chart for average life_expectancy statistic of Swaziland compared to Sub-saharan region
# last_row = len(ws['A'])
# chart1 = BarChart()
# chart1.type = "col"
# chart1.style = 10
# chart1.title = "Average life_expectancy (Years)"
# chart1.y_axis.title = 'Avg life_expectancy'
# chart1.x_axis.title = 'Category'

# data = Reference(ws, min_col=3, min_row=1, max_row=last_row, max_col=3)
# cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
# chart1.add_data(data, titles_from_data=True)
# chart1.set_categories(cats)
# chart1.shape = 4
# ws.add_chart(chart1, "E3")

# wb.save('Bar chart showing average life_expectancy statistic of Swaziland compared to Sub-saharan region.xlsx')



# Question 2

# Extracting a subset of columns for analysis on student IDs
second_sudentid = df_studentid.loc [:, ['Country Name', 'Fertility', 'Life Expectancy', 'YearR']]
#print(second_sudentid)


#Creating new column (decade  1960's to the 2010's)
second_sudentid.loc[((second_sudentid['YearR']>= 1960) & (second_sudentid['YearR'] < 1970)), 'Decade'] = "1960's"
second_sudentid.loc[((second_sudentid['YearR']>= 1970) & (second_sudentid['YearR'] < 1980)), 'Decade'] = "1970's"
second_sudentid.loc[((second_sudentid['YearR']>= 1980) & (second_sudentid['YearR'] < 1990)), 'Decade'] = "1980's"
second_sudentid.loc[((second_sudentid['YearR']>= 1990) & (second_sudentid['YearR'] < 2000)), 'Decade'] = "1990's"
second_sudentid.loc[((second_sudentid['YearR']>= 2000) & (second_sudentid['YearR'] < 2010)), 'Decade'] = "2000's"
second_sudentid.loc[(second_sudentid['YearR']>= 2010), 'Decade'] = "2010's"

#Calculating the average life expectancy by decade for the Second_sudentid DataFrame
third_sudentid = second_sudentid.groupby (['Decade']) [['Life Expectancy']].mean().reset_index()
#print(third_sudentid)



# Creating Line Chart for Expectancy Over Decades in Excel
# from openpyxl.chart import (
#     LineChart,
#     Reference,
# )

# # Append data from the DataFrame to the worksheet
# for r in dataframe_to_rows(second_sudentid, index=False, header=True):
#     ws.append(r)

# # Create a LineChart object
# c1 = LineChart()
# c1.title = "Swaziland Life Expectancy Over the Decades"
# c1.style = 13
# c1.y_axis.title = 'Expectancy'
# c1.x_axis.title = 'Decades'

# # Define data and categories for the chart
# data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
# cats = Reference(ws, min_col=1, min_row=2, max_row=7)

# # Add data and set categories for the chart
# c1.add_data(data, titles_from_data=True)
# c1.set_categories(cats)

# # Add the chart to the worksheet at cell A10
# ws.add_chart(c1, "A10")

# # Save the workbook with the chart
# wb.save('Swaziland average life expectancy over the decades.xlsx')


# Visualizing the correlation between Fertility and Life Expectancy using a scatter plot
plt.scatter(second_sudentid['Fertility'], second_sudentid['Life Expectancy'])
plt.title("Fertility Vs Expectancy") 
plt.xlabel("Fertility (%)")
plt.ylabel("Expectancy")
plt.savefig('Scatter plot showing the correlation between fertility and life expectancy')
plt.show()