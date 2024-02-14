import pandas as pd
from openpyxl import load_workbook
from itertools import islice

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Series, Reference

# Creating a Function to load each excel sheet into a dataframe
def function(origin,datafile):
    wb = load_workbook(filename = origin)
    ws = wb [datafile]
    dataframe = pd.DataFrame(ws.values)
    return dataframe

# Applying the function 
df_metadata = function('week 9/P11-Country-Metadata (1).xlsx', 'Metadata - Countries')
df_fertility = function('week 9/P11-Fertility-Rate.xlsx', 'Data')
df_life_expectancy = function('week 9/P11-Life-Expectancy-At-Birth.xlsx', 'Data' )

#print(df_metadata.head())
#print(df_fertility.head)
#print(df_life_expectancy.head)

#Function to Create worksheets from my workbooks
def Excel(workbooks, worksheets):
    wbb = load_workbook(filename = workbooks)
    wss = wbb [worksheets]
    exl_worksheet = pd.read_excel(workbooks)
    return exl_worksheet

df_metadata3 = Excel('week 9/P11-Country-Metadata (1).xlsx','Metadata - Countries')
df_fertility3 = Excel('week 9/P11-Fertility-Rate.xlsx','Data')
df_life_expectancy3 = Excel( 'week 9/P11-Life-Expectancy-At-Birth.xlsx','Data',)

#print(df_metadata3.head())
#print(df_fertility3.head())
#print(df_life_expectancy3.head())



#   Melting the Year and value(statistic) for Fertility and Life_Expectance datasets
df_cols = ["Country Name", "Country Code", "Indicator Name", "Indicator Code"]
melted_fertility = pd.melt(df_fertility3, id_vars= df_cols, var_name="Year", value_name = "Fertility")


df_cols = ["Country Name", "Country Code", "Indicator Name", "Indicator Code"]
melted_life_expectancy = pd.melt(df_life_expectancy3, id_vars= df_cols, var_name="Year", value_name = "Life Expectancy")


#print(melted_fertility)
#print(melted_life_expectancy)

# Verifying the datatypes of the clolumns in the the dataset
#print(melted_fertility.info())
#print(melted_life_expectancy.info())


# Converting the datatype of the Year column from Object to Integer for both fertility and life_expectancy datasets
melted_fertility ['Year'] = melted_fertility ['Year'].astype(int)
melted_life_expectancy ['Year'] = melted_life_expectancy ['Year'].astype(int)

#print(melted_fertility.info())
#print(melted_life_expectancy.info())

# To display the maximum number of columns to display when showing a DataFrame.
pd.set_option('display.max_columns',500)


# Merging the first two datasets (fertility and metadata) together
df_first_merge = pd.merge (left=df_metadata3,right=melted_fertility,how='left',on=['Country Code'],suffixes = ['L', 'R'])

# Merging the the newly comined datasets (fertility and metadata) with life_expectancy
df_second_merge = pd.merge (left=df_first_merge,right=melted_life_expectancy,how='left',on=['Country Code'],suffixes = ['L', 'R'])

#Analysis
# Creating a dataframe for my StudentID (This filters and shows only the records applicable to 100871852)
df_studentid = df_second_merge[df_second_merge['StudentID'] == 100871852]
#print(df_studentid)


#Average fertility and life expectancy rate for Swaziland over the years
df_country = df_studentid.groupby(['Country Name'])[['Fertility', 'Life Expectancy']].mean().reset_index()
# print(df_country)

# Grouping the data by region to analyze fertility and life expectancy trends(groupby)
df_region = df_second_merge.groupby(['Region'])[['Fertility','Life Expectancy']].mean().reset_index()
df_region = df_region.iloc[[6]]
#print(df_region)

# Creating a common column name for both dataframes (Region and country) so as to perform a join 
df_country.columns = ['Location', 'Fertility', 'Life Expectancy']
df_region.columns = ['Location', 'Fertility', 'Life Expectancy'] 

# print(df_region)
# print(df_country)


# Joining the dataframes region and country together to 
df_join = pd.concat ([df_region, df_country], axis=0)
print(df_join)

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(df_join, index=False, header=True):
    ws.append(r)


last_row = len(ws['A'])
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Average Fertility Rate"
chart1.y_axis.title = 'Avg Fertility'
chart1.x_axis.title = 'Category'

data = Reference(ws, min_col=2, min_row=1, max_row=last_row, max_col=2)
cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, "E3")

wb.save('Sample.xlsx')








#print(df_first_merge)
#print(df_second_merge)













    


# print dataframe



# print(df_metadata.head())
# print("==" * 50)
# print(df_fertility.head())
# print("==" * 50)
# print(df_life_expectancy.head())
# print("==" * 50)
