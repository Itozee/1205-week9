import pandas as pd
from openpyxl import load_workbook
from itertools import islice

#Workbook For metadata
wb_metadata = load_workbook(filename = 'week 9/P11-Country-Metadata.xlsx')

#Worksheet For metadata
ws_metadata = wb_metadata ['Metadata - Countries']

#Dataframe for metadata
df_metadata = pd.DataFrame(ws_metadata.values)
#==============================================================================================
#Workbook for Fertility
wb_fertility = load_workbook(filename = 'week 9/P11-Fertility-Rate.xlsx')

#Worksheet for fertility
ws_fertility = wb_fertility ['Data']

#Dataframe for fertility
df_fertility = pd.DataFrame(ws_fertility.values)
#================================================================================================
#Workbook for life_expectancy
wb_life_expectancy = load_workbook(filename = 'week 9/P11-Life-Expectancy-At-Birth.xlsx')

#Worksheet for life_expectancy
ws_life_expectancy = wb_life_expectancy ['Data']

# Dataframe for life_expectancy
df_life_expectancy = pd.DataFrame(ws_life_expectancy.values)

# print dataframe
#print(df_metadata.head())
print("==" * 50)
#print(df_fertility.head())
print("==" * 50)
#print(df_life_expectancy.head())
print("==" * 50)

# How would you have created a function for all the dataframe created
#=======================================================================================

#To remove the top row index for metadata
data = ws_metadata.values
cols = next(data)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df = pd.DataFrame(data, index=idx, columns=cols)

#print(df.head())

#=======================================================================================

#To remove the top row index for fertility
data = ws_fertility.values
cols = next(data)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df_fertility = pd.DataFrame(data, index=idx, columns=cols)

#print(df_fertility)

#=======================================================================================
#To remove the top row index for life_expectancy
data = ws_life_expectancy.values
cols = next(data)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df = pd.DataFrame(data, index=idx, columns=cols)

print(df.head())
#=======================================================================================

# Create an index with reset_index metadata & skip the first row with iloc
df_metadata = df_metadata.reset_index()
#print(df_metadata)
df_metadata = df_metadata.iloc[1:, :]

# Create an index with reset_index fertility & skip the first row with iloc
df_fertility = df_fertility.reset_index()
#print(df_fertility)
df_fertility = df_fertility.iloc[1:, :]

# Create an index with reset_index life_expectancy & skip the first row with iloc
df_life_expectancy = df_life_expectancy.reset_index()
#print(df_life_expectancy)
df_life_expectancy = df_life_expectancy.iloc[1:, :]

#1
#print(list(df_life_expectancy.columns))

#2
#df_life_expectancy.columns = 
#print(list(df_fertility.columns))

#df_merged = pd.merge(left = df_fertility, right = df_metadata, how = 'left', on=['Country Code'])

#df_merged = pd.merge(left=df_fertility, right=df_metadata, how='left', on=['Country Code'])


#print(df_metadata.head())




