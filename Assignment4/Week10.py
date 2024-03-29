import pandas as pd
from openpyxl import load_workbook
from itertools import islice

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Series, Reference

pd.set_option('display.max_columns', 500)

wb_metadata = load_workbook(filename='Assignment4/P11-Country-Metadata.xlsx' )
ws_metadata = wb_metadata['Metadata - Countries']

wb_fertility = load_workbook(filename='Assignment4/P11-Fertility-Rate.xlsx' )
ws_fertility = wb_fertility['Data']

wb_expectancy = load_workbook(filename='Assignment4/P11-Life-Expectancy-At-Birth.xlsx' )
ws_expectancy = wb_expectancy['Data']

wb_population = load_workbook(filename='Assignment4/P11-Country-Population.xlsx' )
ws_population = wb_population['Data']


# # Creating dataframe metadata
data = ws_metadata.values
cols = next(ws_metadata.values)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df_metadata = pd.DataFrame(data, index=idx, columns=cols)

df_metadata = df_metadata.reset_index()

df_metadata = df_metadata.iloc[1:, :]

# print(df_metadata.head())

# Create dataframe fertility
data = ws_fertility.values
cols = next(ws_fertility.values)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df_fertility = pd.DataFrame(data, index=idx, columns=cols)

df_fertility = df_fertility.reset_index()

df_fertility = df_fertility.iloc[1:, :]

# print(df_fertility.head())

# Create dataframe expectancy
data = ws_expectancy.values
cols = next(ws_expectancy.values)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df_expectancy = pd.DataFrame(data, index=idx, columns=cols)

df_expectancy = df_expectancy.reset_index()

df_expectancy = df_expectancy.iloc[1:, :]

# print(df_expectancy.head())

#  Creating dataframe population
data = ws_population.values
cols = next(ws_population.values)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df_population = pd.DataFrame(data, index=idx, columns=cols)

df_population = df_population.reset_index()

df_population = df_population.iloc[1:, :]



df_metadata.columns = ['Country Name', 'Country Code', 'StudentID', 'Region', 'IncomeGroup', 'SpecialNotes']
df_fertility.columns = ['Country Name', 'Country Code', 'Indicator Name', 'Indicator Code', '1960', '1961', '1962', '1963',
'1964', '1965', '1966', '1967', '1968', '1969', '1970', '1971', '1972', '1973', '1974', '1975',
'1976', '1977', '1978', '1979', '1980', '1981', '1982', '1983', '1984', '1985', '1986', '1987',
'1988', '1989', '1990', '1991', '1992', '1993', '1994', '1995', '1996', '1997', '1998', '1999',
'2000', '2001', '2002', '2003', '2004', '2005', '2006', '2007', '2008', '2009', '2010', '2011', '2012', '2013']
df_expectancy.columns = ['Country Name', 'Country Code', 'Indicator Name', 'Indicator Code', '1960', '1961', '1962', '1963',
'1964', '1965', '1966', '1967', '1968', '1969', '1970', '1971', '1972', '1973', '1974', '1975', '1976', '1977',
'1978', '1979', '1980', '1981', '1982', '1983', '1984', '1985', '1986', '1987', '1988', '1989', '1990', '1991', '1992',
'1993', '1994', '1995', '1996', '1997', '1998', '1999', '2000', '2001', '2002', '2003', '2004', '2005', '2006', '2007', '2008', '2009', '2010', '2011', '2012', '2013']
df_population.columns = ['Country Name', 'Country Code', 'Indicator Name', 'Indicator Code', '1960', '1961', '1962', '1963', '1964', '1965', '1966', '1967', '1968', '1969',
'1970', '1971', '1972', '1973', '1974', '1975', '1976', '1977', '1978', '1979', '1980', '1981', '1982', '1983', '1984', '1985', '1986', '1987', '1988', '1989', '1990', '1991', '1992', '1993', '1994', '1995', '1996', '1997', '1998', '1999', '2000', '2001', '2002', '2003', '2004', '2005', '2006', '2007', '2008', '2009', '2010', '2011', '2012', '2013']




#Melting Life_expectancy, Fertility and Population DataFrame
# Defining the columns that wount be melted

df_fertility_cols = ["Country Name", "Country Code", "Indicator Name", "Indicator Code"]
melted_fertility = pd.melt(df_fertility, id_vars= df_fertility_cols, var_name="Year", value_name = "Fertility")

df_expectancy_cols = ["Country Name", "Country Code", "Indicator Name", "Indicator Code"]
melted_expectancy = pd.melt(df_expectancy, id_vars= df_expectancy_cols, var_name="Year", value_name = "Expectancy")

df_population_cols = ["Country Name", "Country Code", "Indicator Name", "Indicator Code"]
melted_population = pd.melt(df_population, id_vars= df_population_cols, var_name="Year", value_name = "Population")

# Converting the Year to integers.
melted_expectancy['Year'] = melted_expectancy['Year'].astype(int)
melted_fertility['Year'] = melted_fertility['Year'].astype(int)

# Merging fertility with metadata
df_merged = pd.merge(left=melted_fertility, right=df_metadata, how='left', on=['Country Code'], suffixes= ['_f' , '_m'] )
# print(df_merged)

# Merging df_merged with df_expectancy
df_mergedd = pd.merge(left=df_merged, right=melted_expectancy, how='left', on=['Country Code'], suffixes= ['_fertility', '_expectancy'])

# # Merging df_merged with df_population
# df_mergeddd = pd.merge(left=df_mergedd, right=melted_population, how='left', on=['Country Code'] )
# print(df_mergeddd)

# Creating a dataframe to capture the Regions Fertility and Life Expectancy
df_region = df_mergedd.groupby(['Region'])[['Fertility', 'Expectancy']].mean().reset_index()
df_region = df_region.iloc[[6]]
print(df_region)

# Creating a dataframe for my StudentID
df_Country = df_mergedd[df_mergedd['StudentID'] == 100871852]
df_Countryy = df_Country.loc[:, ['Country Name', 'Fertility', 'Expectancy','Year_expectancy']]
df_Country_A = df_Country.groupby(['Country Name'])[['Fertility', 'Expectancy']].mean().reset_index()

#print(df_Country_A)

# Creating a common column name
df_region.columns = ['Location', 'Fertility', 'Expectancy']
df_Country_A.columns = ['Location', 'Fertility', 'Expectancy']

# merging data frames
df_join = pd.concat([df_region, df_Country_A], axis =0 )

# Plotting bar chart to compare Sub-Saharan Africa and Gambia Average Fertility
#  and Average Life Expectancy

wb = Workbook()
ws = wb.active

# for r in dataframe_to_rows(df_join, index=False, header=True):
#     ws.append(r)


# last_row = len(ws['A'])
# chart1 = BarChart()
# chart1.type = "col"
# chart1.style = 10
# chart1.title = "Average Fertility Rate"
# chart1.y_axis.title = 'Avg Fertility'
# chart1.x_axis.title = 'Category'

# data = Reference(ws, min_col=2, min_row=1, max_row=last_row, max_col=2)
# cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
# chart1.add_data(data, titles_from_data=True)
# chart1.set_categories(cats)
# chart1.shape = 4
# ws.add_chart(chart1, "E3")

# wb.save('Sample.xlsx')



# Creating new column (decade  1960's to the 2010's)

df_Countryy.loc[((df_Countryy['Year_expectancy']>= 1960) & (df_Countryy['Year_expectancy'] < 1970)), 'Decade'] = "1960's"
df_Countryy.loc[((df_Countryy['Year_expectancy']> 1970) & (df_Countryy['Year_expectancy'] < 1980)), 'Decade'] = "1970's"
df_Countryy.loc[((df_Countryy['Year_expectancy']> 1980) & (df_Countryy['Year_expectancy'] < 1990)), 'Decade'] = "1980's"
df_Countryy.loc[((df_Countryy['Year_expectancy']> 1990) & (df_Countryy['Year_expectancy'] < 2000)), 'Decade'] = "1990's"
df_Countryy.loc[((df_Countryy['Year_expectancy']> 2000) & (df_Countryy['Year_expectancy'] < 2010)), 'Decade'] = "2000's"
df_Countryy.loc[(df_Countryy['Year_expectancy']>= 2010), 'Decade'] = "2010's"


# Creating a dataframe grouping by decade to get average Life_Expectancy and Fertility
#df_mergedd_D = df_mergedd.groupby(['Decade'])[['Expectancy']].mean().reset_index()

#df_mergedd_DA = df_mergedd.groupby(['Decade'])[['Expectancy', 'Fertility']].mean().reset_index()

# print(df_mergedd_DA)

df_Countryy = df_Countryy.groupby(['Decade'])[['Expectancy']].mean().reset_index()

# df_Countryy = df_Countryy.groupby(['Decade'])[['Expectancy', 'Fertility']].mean().reset_index()
#print(df_Countryy)

from openpyxl.chart import (
    LineChart,
    Reference,
)

for r in dataframe_to_rows(df_Countryy, index=False, header=True):
    ws.append(r)

c1 = LineChart()
c1.title = "Expectancy Over the Decades"
c1.style = 13
c1.y_axis.title = 'Expectancy'
c1.x_axis.title = 'Decades'

data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)

cats = Reference(ws, min_col=1, min_row=2, max_row=7)
c1.add_data(data, titles_from_data=True)
c1.set_categories(cats)


ws.add_chart(c1, "A10")

wb.save('Sample_5.xlsx')


# Scatter plot to check for relationship between Fertility and Expectancy
# import matplotlib.pyplot as plt
# plt.scatter(df_mergedd_DA.Fertility, df_mergedd_DA.Expectancy)
# plt.title("Fertility Vs Expectancy")
# plt.xlabel("Fertility (%)")
# plt.ylabel("Expectancy")
# plt.show()
# plt.savefig('sample_3')
